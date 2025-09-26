#!/usr/bin/env python3
"""
Battery Capacity Sweep Results Aggregator

Reads individual CSV files from battery capacity sweeps and creates
consolidated comparison files for A1 and A2 scenarios, similar to the
client deliverable format.
"""

import pandas as pd
import numpy as np
from pathlib import Path
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def read_and_categorize_csv(file_path, solar_kw, load_scenario, scenario_suffix):
    """
    Read a CSV file and add categorization columns.
    
    Args:
        file_path (str): Path to the CSV file
        solar_kw (float): Solar system size in kW DC
        load_scenario (str): Load scenario (High, Med, Low)
        scenario_suffix (str): Scenario suffix (A1 or A2)
    
    Returns:
        pd.DataFrame: DataFrame with added categorization columns
    """
    try:
        df = pd.read_csv(file_path)
        
        # Add categorization columns
        df['Solar_kw_DC'] = solar_kw
        df['EV_load_scenario'] = load_scenario
        df['Scenario'] = scenario_suffix
        
        # Ensure consistent column naming
        if 'capacity_kwh' not in df.columns and 'Capacity (kWh)' in df.columns:
            df = df.rename(columns={'Capacity (kWh)': 'capacity_kwh'})
        
        # Round values for consistency
        numeric_columns = df.select_dtypes(include=[np.number]).columns
        df[numeric_columns] = df[numeric_columns].round(2)
        
        return df
    
    except FileNotFoundError:
        print(f"Warning: File {file_path} not found, skipping...")
        return pd.DataFrame()
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return pd.DataFrame()


def aggregate_results(scenario_suffix='A1'):
    """
    Aggregate all battery sweep results for a given scenario.
    
    Args:
        scenario_suffix (str): Either 'A1' or 'A2'
    
    Returns:
        pd.DataFrame: Consolidated results
    """
    results = []
    
    # Define file mapping: (filename, solar_kw, load_scenario)
    file_mappings = [
        # 259kW PV system (Large PV)
        (f'comparison_med_large_pv_{scenario_suffix.lower()}.csv', 259.2, 'Med'),
        (f'comparison_low_large_pv_{scenario_suffix.lower()}.csv', 259.2, 'Low'),
        
        # 194kW PV system (Medium PV)  
        (f'comparison_high_medium_pv_{scenario_suffix.lower()}.csv', 194.4, 'High'),
        (f'comparison_med_medium_pv_{scenario_suffix.lower()}.csv', 194.4, 'Med'),
        (f'comparison_low_medium_pv_{scenario_suffix.lower()}.csv', 194.4, 'Low'),
    ]
    
    for filename, solar_kw, load_scenario in file_mappings:
        print(f"Processing {filename}...")
        df = read_and_categorize_csv(filename, solar_kw, load_scenario, scenario_suffix)
        if not df.empty:
            results.append(df)
    
    if not results:
        print(f"No valid data found for scenario {scenario_suffix}")
        return pd.DataFrame()
    
    # Combine all results
    combined_df = pd.concat(results, ignore_index=True)
    
    # Sort by solar size, load scenario, and capacity for consistent ordering
    load_order = {'Low': 1, 'Med': 2, 'High': 3}
    combined_df['load_order'] = combined_df['EV_load_scenario'].map(load_order)
    combined_df = combined_df.sort_values(['Solar_kw_DC', 'load_order', 'capacity_kwh'])
    combined_df = combined_df.drop('load_order', axis=1)
    
    return combined_df


def format_excel_output(df, output_filename):
    """
    Format the DataFrame and save as Excel with styling similar to the template.
    
    Args:
        df (pd.DataFrame): Data to save
        output_filename (str): Output Excel filename
    """
    # Create a new workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Full comparison"
    
    # Define column order to match template
    column_order = [
        'capacity_kwh', 'power_kw', 'Solar_kw_DC', 'EV_load_scenario',
        'total_load_kwh', 'total_pv_kwh', 'pv_curtailed_kwh', 'pv_used_kwh',
        'energy_not_served_kwh', 'self_sufficiency_percent', 'pv_utilization_percent',
        'throughput_charge_kwh', 'throughput_discharge_kwh', 'estimated_full_cycles',
        'loss_of_load_hours', 'max_contiguous_lol_hours'
    ]
    
    # Select and reorder columns
    available_columns = [col for col in column_order if col in df.columns]
    if available_columns:
        df_output = df[available_columns].copy()
    else:
        df_output = df.copy()
    
    # Add the DataFrame to the worksheet
    for r in dataframe_to_rows(df_output, index=False, header=True):
        ws.append(r)
    
    # Format header row
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(output_filename)
    print(f"Excel file saved: {output_filename}")


def main():
    """Main function to aggregate battery sweep results."""
    parser = argparse.ArgumentParser(description='Aggregate battery capacity sweep results')
    parser.add_argument('--output-dir', default='.', help='Output directory for result files')
    parser.add_argument('--format', choices=['csv', 'excel', 'both'], default='both', 
                       help='Output format')
    
    args = parser.parse_args()
    
    output_dir = Path(args.output_dir)
    output_dir.mkdir(exist_ok=True)
    
    # Process both A1 and A2 scenarios
    for scenario in ['A1', 'A2']:
        print(f"\n{'='*50}")
        print(f"Processing Scenario {scenario}")
        print(f"{'='*50}")
        
        # Aggregate results
        combined_df = aggregate_results(scenario)
        
        if combined_df.empty:
            print(f"No data found for scenario {scenario}, skipping...")
            continue
        
        print(f"Combined {len(combined_df)} rows for scenario {scenario}")
        print(f"Solar systems: {sorted(combined_df['Solar_kw_DC'].unique())} kW")
        print(f"Load scenarios: {sorted(combined_df['EV_load_scenario'].unique())}")
        print(f"Capacity range: {combined_df['capacity_kwh'].min()} - {combined_df['capacity_kwh'].max()} kWh")
        
        # Generate output filenames
        base_filename = f"SGWS_Scenario_comparison_{scenario}"
        csv_filename = output_dir / f"{base_filename}.csv"
        excel_filename = output_dir / f"{base_filename}.xlsx"
        
        # Save outputs
        if args.format in ['csv', 'both']:
            combined_df.to_csv(csv_filename, index=False)
            print(f"CSV file saved: {csv_filename}")
        
        if args.format in ['excel', 'both']:
            format_excel_output(combined_df, excel_filename)
        
        # Display summary statistics
        print(f"\nSummary for {scenario}:")
        print(f"  Total scenarios: {len(combined_df.groupby(['Solar_kw_DC', 'EV_load_scenario']))}")
        if 'self_sufficiency_percent' in combined_df.columns:
            print(f"  Self-sufficiency range: {combined_df['self_sufficiency_percent'].min():.1f}% - {combined_df['self_sufficiency_percent'].max():.1f}%")
        if 'pv_utilization_percent' in combined_df.columns:
            print(f"  PV utilization range: {combined_df['pv_utilization_percent'].min():.1f}% - {combined_df['pv_utilization_percent'].max():.1f}%")
    
    print(f"\n{'='*50}")
    print("Aggregation completed!")
    print("Files generated:")
    print("  - SGWS_Scenario_comparison_A1.xlsx")
    print("  - SGWS_Scenario_comparison_A2.xlsx")
    if args.format in ['csv', 'both']:
        print("  - SGWS_Scenario_comparison_A1.csv")
        print("  - SGWS_Scenario_comparison_A2.csv")


if __name__ == "__main__":
    main()
